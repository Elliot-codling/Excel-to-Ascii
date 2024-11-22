#Convert Excell art to code blocks for Ascii Invaders

import openpyxl

textContents = []

wb = openpyxl.load_workbook('Ascii ArtWork.xlsx',data_only=True)
fs = wb.active
fs_count_row = fs.max_row 
fs_count_col = fs.max_column 

for row in range(1,fs_count_row+1):
    for column in range(1,fs_count_col+1):
        cell_color = fs.cell(column=column, row=row)
        fgColor = cell_color.fill.fgColor.index
        if fgColor=='00000000':
            continue
        else:
            textContents += [fgColor]

print(textContents)
print(f"Number of columns: ", fs_count_col)
print(f"Number of rows: ", fs_count_row)
#Write over previous contents
file = open("Text.txt", "w")
file.write("")
file.close()

index = 0
try:
    file = open("Text.txt", "a")
    for row in range(fs_count_row):
        for col in range(fs_count_col):
            if textContents[index] == 1:
                file.write("{' ', BACKGROUND_BLACK},")
            else:
                file.write("{' ', BACKGROUND_WHITE},")
            index += 1
        file.write("\n")
except:
    pass

file.close()